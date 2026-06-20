"""
Professional Excel export module for Financial Modeler.

Generates multi-sheet workbook with:
- Assumptions
- Revenue
- Payroll
- Operating Expenses
- Financing
- Profit & Loss
- Cash Flow

With professional formatting:
- Bold headers
- Currency formatting
- Percentage formatting
- Frozen header rows
- Auto-sized columns
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import Optional
from config.version import PLATFORM_VERSION
from engine.model import build_model

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_excel_workbook(model_inputs: dict, scenario_name: str = "Business Scenario") -> bytes:
    """
    Generate professional Excel workbook with all model data.
    
    Args:
        model_inputs: Dictionary of model inputs from session state
        scenario_name: Name of the business/scenario
    
    Returns:
        bytes: Excel file as bytes for download
    
    Raises:
        ImportError: If openpyxl is not installed
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )
    
    # Build model to get calculated outputs
    try:
        outputs = build_model(model_inputs)
    except Exception as e:
        # If model build fails, create workbook with inputs only
        outputs = None
    
    buffer = BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Sheet 1: Assumptions (with metadata)
        _write_assumptions_sheet(writer, model_inputs, scenario_name)
        
        # Sheet 2: Revenue
        _write_revenue_sheet(writer, model_inputs, outputs)
        
        # Sheet 3: Payroll
        _write_payroll_sheet(writer, model_inputs)
        
        # Sheet 4: Operating Expenses
        _write_opex_sheet(writer, model_inputs)
        
        # Sheet 5: Financing
        _write_financing_sheet(writer, model_inputs, outputs)
        
        # Sheet 6: Profit & Loss
        if outputs and 'income_statement' in outputs:
            _write_profit_loss_sheet(writer, outputs['income_statement'])
        
        # Sheet 7: Cash Flow
        if outputs and 'cash_flow' in outputs:
            _write_cash_flow_sheet(writer, outputs['cash_flow'])
    
    buffer.seek(0)
    return buffer.getvalue()


def _write_assumptions_sheet(writer, model_inputs: dict, scenario_name: str):
    """Write Assumptions sheet with metadata and key settings."""
    
    # Metadata section
    metadata = [
        ['FINANCIAL MODEL - ASSUMPTIONS'],
        [''],
        ['Business Name', scenario_name],
        ['Generated Date/Time', datetime.now().strftime('%Y-%m-%d %I:%M %p')],
        ['Platform Version', PLATFORM_VERSION],
        ['Forecast Length', f"{model_inputs.get('periods', 36)} periods"],
        ['Time Mode', model_inputs.get('time_mode', 'monthly').capitalize()],
        [''],
        ['MODEL SETTINGS'],
        [''],
        ['Revenue Settings', ''],
        ['  Global COGS %', model_inputs.get('global_cogs_pct', 0.30)],
        ['  COGS Improvement per Year %', model_inputs.get('cogs_improvement_pct', 0.0)],
        ['  Startup Ramp (Months)', model_inputs.get('startup_ramp_months', 0)],
        [''],
        ['Financing Settings', ''],
        ['  Loan Principal', model_inputs.get('loan_principal', 50000.0)],
        ['  Annual Interest Rate', model_inputs.get('loan_annual_rate', 0.06)],
        ['  Loan Term (Months)', model_inputs.get('loan_term_months', 60)],
        [''],
        ['Working Capital Settings', ''],
        ['  Accounts Receivable (Days)', model_inputs.get('ar_days', 0)],
        ['  Accounts Payable (Days)', model_inputs.get('ap_days', 0)],
        ['  Inventory (Days)', model_inputs.get('inventory_days', 0)],
        [''],
        ['Owner Compensation', ''],
        ['  Mode', model_inputs.get('owner_compensation_mode', 'payroll').capitalize()],
        ['  Annual Amount', model_inputs.get('owner_compensation_annual', 0.0)],
    ]
    
    # Add seasonality if enabled
    seasonality = model_inputs.get('seasonality', {})
    if seasonality.get('enabled'):
        metadata.append([''])
        metadata.append(['Seasonality', ''])
        metadata.append(['  Mode', seasonality.get('mode', 'OFF')])
    
    # Create DataFrame
    df = pd.DataFrame(metadata, columns=['Parameter', 'Value'])
    df.to_excel(writer, sheet_name='Assumptions', index=False, header=False)
    
    # Format the sheet
    worksheet = writer.sheets['Assumptions']
    
    # Title formatting
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    
    # Section headers (rows with all caps text)
    for row_idx, row in enumerate(metadata, start=1):
        if row[0] and row[0].isupper() and row[0] != 'FINANCIAL MODEL - ASSUMPTIONS':
            cell = worksheet[f'A{row_idx}']
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Format percentage values
    for row_idx, row in enumerate(metadata, start=1):
        if len(row) > 1 and row[1] is not None:
            cell = worksheet[f'B{row_idx}']
            if 'Rate' in str(row[0]) or '%' in str(row[0]):
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
            elif 'Amount' in str(row[0]) or 'Principal' in str(row[0]):
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    # Column widths
    worksheet.column_dimensions['A'].width = 35
    worksheet.column_dimensions['B'].width = 25
    
    # Freeze top row
    worksheet.freeze_panes = 'A2'


def _write_revenue_sheet(writer, model_inputs: dict, outputs: Optional[dict]):
    """Write Revenue sheet with stream details and forecast."""
    
    revenue_streams = model_inputs.get('revenue_streams', [])
    
    if not revenue_streams:
        # Empty sheet with message
        df = pd.DataFrame([['No revenue streams configured']], columns=['Message'])
        df.to_excel(writer, sheet_name='Revenue', index=False)
        return
    
    # Build revenue streams table
    rows = []
    for i, stream in enumerate(revenue_streams, 1):
        rows.append({
            'Stream #': i,
            'Name': stream.get('name', 'Unnamed'),
            'Price per Unit': stream.get('price', 0.0),
            'Initial Volume': stream.get('volume', 0.0),
            'Growth Rate': stream.get('growth_rate', 0.0),
            'COGS %': stream.get('cogs_override') if stream.get('cogs_override') is not None else model_inputs.get('global_cogs_pct', 0.30)
        })
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Revenue', index=False, startrow=0)
    
    # Format the sheet
    worksheet = writer.sheets['Revenue']
    
    # Header formatting
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Format currency columns
    for row_num in range(2, len(df) + 2):
        worksheet.cell(row=row_num, column=3).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Price
        worksheet.cell(row=row_num, column=4).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1  # Volume
        worksheet.cell(row=row_num, column=5).number_format = numbers.FORMAT_PERCENTAGE_00  # Growth Rate
        worksheet.cell(row=row_num, column=6).number_format = numbers.FORMAT_PERCENTAGE_00  # COGS %
    
    # Auto-size columns
    for col_num in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col_num)
        max_length = max(
            len(str(df.columns[col_num - 1])),
            max(len(str(val)) for val in df.iloc[:, col_num - 1]) if len(df) > 0 else 0
        )
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Add revenue forecast if available
    if outputs and 'revenue_df' in outputs:
        revenue_df = outputs['revenue_df']
        
        # Add spacing
        start_row = len(df) + 4
        
        # Add forecast section header
        worksheet.cell(row=start_row, column=1).value = 'REVENUE FORECAST BY PERIOD'
        worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        
        # Write forecast data
        revenue_df.to_excel(writer, sheet_name='Revenue', index=True, startrow=start_row + 1, header=True)
        
        # Format forecast headers
        for col_num in range(1, len(revenue_df.columns) + 2):
            cell = worksheet.cell(row=start_row + 2, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')


def _write_payroll_sheet(writer, model_inputs: dict):
    """Write Payroll sheet with role details."""
    
    payroll_roles = model_inputs.get('payroll_roles', [])
    
    if not payroll_roles:
        df = pd.DataFrame([['No payroll roles configured']], columns=['Message'])
        df.to_excel(writer, sheet_name='Payroll', index=False)
        return
    
    # Build payroll table
    rows = []
    for i, role in enumerate(payroll_roles, 1):
        pay_type = role.get('pay_type', 'salary')
        
        if pay_type == 'salary':
            rate_label = 'Annual Salary'
            rate_value = role.get('rate', 0.0)
        else:
            rate_label = 'Hourly Rate'
            rate_value = role.get('rate', 0.0)
        
        rows.append({
            'Role #': i,
            'Role Name': role.get('role', 'Unnamed'),
            'Headcount': role.get('headcount', 1),
            'Pay Type': pay_type.capitalize(),
            rate_label: rate_value,
            'Hours per Week': role.get('hours_per_week', 40) if pay_type == 'hourly' else 'N/A',
            'Annual Raise %': role.get('annual_raise_pct', 0.0),
            'Payroll Tax %': role.get('payroll_tax_pct', 0.0765),
            'Benefits %': role.get('benefits_pct', 0.15),
            'Role Type': role.get('role_type', 'indirect').capitalize()
        })
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Payroll', index=False)
    
    # Format the sheet
    worksheet = writer.sheets['Payroll']
    
    # Header formatting
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Format columns
    for row_num in range(2, len(df) + 2):
        worksheet.cell(row=row_num, column=5).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Rate
        worksheet.cell(row=row_num, column=7).number_format = numbers.FORMAT_PERCENTAGE_00  # Raise %
        worksheet.cell(row=row_num, column=8).number_format = numbers.FORMAT_PERCENTAGE_00  # Tax %
        worksheet.cell(row=row_num, column=9).number_format = numbers.FORMAT_PERCENTAGE_00  # Benefits %
    
    # Auto-size columns
    for col_num in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 15
    
    worksheet.column_dimensions['B'].width = 25  # Role Name
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def _write_opex_sheet(writer, model_inputs: dict):
    """Write Operating Expenses sheet."""
    
    opex_items = model_inputs.get('opex_items', [])
    
    if not opex_items:
        df = pd.DataFrame([['No operating expenses configured']], columns=['Message'])
        df.to_excel(writer, sheet_name='Operating Expenses', index=False)
        return
    
    # Build opex table
    rows = []
    for i, item in enumerate(opex_items, 1):
        rows.append({
            'Item #': i,
            'Expense Name': item.get('name', 'Unnamed'),
            'Amount': item.get('amount', 0.0),
            'Growth Rate': item.get('growth_rate', 0.03),
            'Category': item.get('category', 'fixed').capitalize()
        })
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Operating Expenses', index=False)
    
    # Format the sheet
    worksheet = writer.sheets['Operating Expenses']
    
    # Header formatting
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Format columns
    for row_num in range(2, len(df) + 2):
        worksheet.cell(row=row_num, column=3).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Amount
        worksheet.cell(row=row_num, column=4).number_format = numbers.FORMAT_PERCENTAGE_00  # Growth Rate
    
    # Auto-size columns
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def _write_financing_sheet(writer, model_inputs: dict, outputs: Optional[dict]):
    """Write Financing sheet with loan details and schedule."""
    
    # Basic loan info
    loan_data = [
        ['LOAN DETAILS', ''],
        ['Loan Principal', model_inputs.get('loan_principal', 50000.0)],
        ['Annual Interest Rate', model_inputs.get('loan_annual_rate', 0.06)],
        ['Loan Term (Months)', model_inputs.get('loan_term_months', 60)],
        ['Loan Start Period', model_inputs.get('loan_start_period', 0)],
    ]
    
    # Advanced financing if in advanced mode
    if model_inputs.get('mode') == 'Advanced':
        financing_sources = model_inputs.get('financing_sources', [])
        if financing_sources:
            loan_data.append([''])
            loan_data.append(['FINANCING SOURCES', ''])
            for source in financing_sources:
                loan_data.append([source.get('name', 'Unnamed'), source.get('amount', 0.0)])
    
    df = pd.DataFrame(loan_data, columns=['Parameter', 'Value'])
    df.to_excel(writer, sheet_name='Financing', index=False, header=False)
    
    # Format the sheet
    worksheet = writer.sheets['Financing']
    
    # Header formatting
    worksheet['A1'].font = Font(bold=True, size=11)
    worksheet['A1'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Format currency and percentage
    for row_idx, row in enumerate(loan_data, start=1):
        if len(row) > 1 and row[1] != '':
            cell = worksheet[f'B{row_idx}']
            if 'Rate' in str(row[0]):
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
            elif 'Principal' in str(row[0]) or 'Amount' in str(row[0]) or isinstance(row[1], (int, float)):
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    # Column widths
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 20
    
    # Add loan schedule if available
    if outputs and 'loan_schedule' in outputs:
        loan_schedule = outputs['loan_schedule']
        
        # Add spacing
        start_row = len(loan_data) + 3
        
        # Add schedule header
        worksheet.cell(row=start_row, column=1).value = 'LOAN AMORTIZATION SCHEDULE'
        worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        
        # Write schedule
        loan_schedule.to_excel(writer, sheet_name='Financing', index=True, startrow=start_row + 1)
        
        # Format schedule headers
        for col_num in range(1, len(loan_schedule.columns) + 2):
            cell = worksheet.cell(row=start_row + 2, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')


def _write_profit_loss_sheet(writer, income_statement: pd.DataFrame):
    """Write Profit & Loss (Income Statement) sheet."""
    
    # Transpose: line items as rows, periods as columns
    income_transposed = income_statement.T
    income_transposed.columns = [f'Period {i}' for i in income_transposed.columns]
    income_transposed.index.name = 'Line Item'
    
    income_transposed.to_excel(writer, sheet_name='Profit & Loss', index=True)
    
    # Format the sheet
    worksheet = writer.sheets['Profit & Loss']
    
    # Header formatting
    for col_num in range(1, len(income_transposed.columns) + 2):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Format all data cells as currency
    for row_num in range(2, len(income_transposed) + 2):
        for col_num in range(2, len(income_transposed.columns) + 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    # Bold line item names
    for row_num in range(2, len(income_transposed) + 2):
        worksheet.cell(row=row_num, column=1).font = Font(bold=True)
    
    # Column widths
    worksheet.column_dimensions['A'].width = 30
    for col_num in range(2, len(income_transposed.columns) + 2):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 15
    
    # Freeze header row and first column
    worksheet.freeze_panes = 'B2'


def _write_cash_flow_sheet(writer, cash_flow: pd.DataFrame):
    """Write Cash Flow sheet."""
    
    # Transpose: line items as rows, periods as columns
    cash_flow_transposed = cash_flow.T
    cash_flow_transposed.columns = [f'Period {i}' for i in cash_flow_transposed.columns]
    cash_flow_transposed.index.name = 'Line Item'
    
    cash_flow_transposed.to_excel(writer, sheet_name='Cash Flow', index=True)
    
    # Format the sheet
    worksheet = writer.sheets['Cash Flow']
    
    # Header formatting
    for col_num in range(1, len(cash_flow_transposed.columns) + 2):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Format all data cells as currency
    for row_num in range(2, len(cash_flow_transposed) + 2):
        for col_num in range(2, len(cash_flow_transposed.columns) + 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    # Bold line item names
    for row_num in range(2, len(cash_flow_transposed) + 2):
        worksheet.cell(row=row_num, column=1).font = Font(bold=True)
    
    # Column widths
    worksheet.column_dimensions['A'].width = 30
    for col_num in range(2, len(cash_flow_transposed.columns) + 2):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 15
    
    # Freeze header row and first column
    worksheet.freeze_panes = 'B2'


def get_excel_filename(scenario_name: str) -> str:
    """
    Generate Excel filename with format:
    {BusinessName}_FinancialModel_YYYYMMDD_HHMM.xlsx
    
    Args:
        scenario_name: Name of the business/scenario
    
    Returns:
        Formatted filename
    """
    # Sanitize business name for filename
    safe_name = "".join(c for c in scenario_name if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_name = safe_name.replace(' ', '_')
    
    if not safe_name:
        safe_name = "FinancialModel"
    
    # Generate timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    
    return f"{safe_name}_FinancialModel_{timestamp}.xlsx"
