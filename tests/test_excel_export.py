"""
Tests for Excel scenario export functionality.
"""
import pytest
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from utils.exporters import export_scenario_to_excel


@pytest.fixture
def sample_model_inputs():
    """Sample model inputs for testing."""
    return {
        'scenario_name': 'Test Scenario',
        'time_mode': 'monthly',
        'periods': 12,
        'global_cogs_pct': 0.30,
        'cogs_improvement_pct': 2.0,
        'startup_ramp_months': 6,
        'revenue_streams': [
            {
                'name': 'Product Sales',
                'price': 100.0,
                'volume': 50.0,
                'growth_rate': 0.10,
                'cogs_override': None
            }
        ],
        'loan_principal': 100000.0,
        'loan_annual_rate': 0.06,
        'loan_term_months': 60,
        'loan_start_period': 0,
        'mode': 'Basic',
        'owner_compensation': {
            'mode': 'distribution',
            'amount': 60000.0
        },
        'tax_rate': 0.25,
        'annual_depreciation': 5000.0,
        'ar_days': 30,
        'ap_days': 30,
        'inventory_days': 15,
        'payroll_roles': [
            {
                'name': 'Manager',
                'count': 1,
                'annual_salary': 50000.0,
                'payroll_tax_rate': 0.15,
                'category': 'indirect'
            }
        ]
    }


@pytest.fixture
def sample_income_statement_df():
    """Sample income statement DataFrame for testing."""
    periods = 12
    data = {
        f'Period {i}': [10000, 3000, 7000, 5000, 1000, 1000, 417, 500, 83, 21, 62]
        for i in range(periods)
    }
    df = pd.DataFrame(data, index=[
        'Revenue', 'COGS', 'Gross Profit', 'Payroll', 'Overhead',
        'EBITDA', 'Depreciation', 'Interest', 'Pre-Tax Income', 'Taxes', 'Net Income'
    ])
    df.index.name = 'Line Item'
    return df


@pytest.fixture
def sample_cash_flow_df():
    """Sample cash flow DataFrame for testing."""
    periods = 12
    ending_cash_values = [5000, 5500, 6000, 6500, 7000, 7500, 8000, 8500, 9000, 9500, 10000, 10500]
    data = {
        f'Period {i}': [1500, 1000, 500, ending_cash_values[i]]
        for i in range(periods)
    }
    df = pd.DataFrame(data, index=[
        'Operating Cash Flow', 'Debt Service', 'Owner Distribution', 'Ending Cash'
    ])
    df.index.name = 'Line Item'
    return df


def test_export_scenario_to_excel_basic(sample_model_inputs):
    """Test basic Excel export without DataFrames."""
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0
    
    # Load workbook to verify structure
    wb = load_workbook(BytesIO(excel_bytes))
    
    # Check sheets exist (only Scenario and Raw JSON without DataFrames)
    assert 'Scenario' in wb.sheetnames
    assert 'Raw_JSON' in wb.sheetnames
    assert 'Summary' not in wb.sheetnames  # No Summary without DataFrames


def test_export_scenario_to_excel_with_dataframes(sample_model_inputs, sample_income_statement_df, sample_cash_flow_df):
    """Test Excel export with DataFrames."""
    excel_bytes = export_scenario_to_excel(
        sample_model_inputs,
        income_statement_df=sample_income_statement_df,
        cash_flow_df=sample_cash_flow_df
    )
    
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0
    
    # Load workbook to verify structure
    wb = load_workbook(BytesIO(excel_bytes))
    
    # Check all sheets exist
    assert 'Summary' in wb.sheetnames
    assert 'Scenario' in wb.sheetnames
    assert 'Income_Statement' in wb.sheetnames
    assert 'Cash_Flow' in wb.sheetnames
    assert 'Raw_JSON' in wb.sheetnames


def test_scenario_sheet_contains_metadata(sample_model_inputs):
    """Test that Scenario sheet contains key metadata."""
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Scenario')
    
    # Check key parameters are present
    parameters = df['Parameter'].tolist()
    
    assert 'Scenario Name' in parameters
    assert 'Generated At' in parameters
    assert 'Time Mode' in parameters
    assert 'Periods' in parameters
    assert 'Global COGS %' in parameters
    assert 'Loan Principal' in parameters


def test_scenario_sheet_values_correct(sample_model_inputs):
    """Test that Scenario sheet values match input data."""
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Scenario')
    
    # Create lookup dictionary
    data_dict = dict(zip(df['Parameter'], df['Value']))
    
    # Verify values
    assert data_dict['Scenario Name'] == 'Test Scenario'
    assert data_dict['Time Mode'] == 'monthly'
    assert data_dict['Periods'] == 12
    assert data_dict['Global COGS %'] == '30.0%'
    assert data_dict['Loan Principal'] == '$100,000.00'


def test_income_statement_sheet_structure(sample_model_inputs, sample_income_statement_df):
    """Test that Income_Statement sheet has correct structure."""
    excel_bytes = export_scenario_to_excel(
        sample_model_inputs,
        income_statement_df=sample_income_statement_df
    )
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Income_Statement', index_col=0)
    
    # Check rows exist
    expected_rows = [
        'Revenue', 'COGS', 'Gross Profit', 'Payroll', 'Overhead',
        'EBITDA', 'Depreciation', 'Interest', 'Pre-Tax Income', 'Taxes', 'Net Income'
    ]
    
    for row in expected_rows:
        assert row in df.index


def test_income_statement_sheet_values(sample_model_inputs, sample_income_statement_df):
    """Test that Income_Statement sheet values match DataFrame."""
    excel_bytes = export_scenario_to_excel(
        sample_model_inputs,
        income_statement_df=sample_income_statement_df
    )
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Income_Statement', index_col=0)
    
    # Verify first period values
    assert df['Period 0']['Revenue'] == 10000
    assert df['Period 0']['COGS'] == 3000
    assert df['Period 0']['Gross Profit'] == 7000
    assert df['Period 0']['Net Income'] == 62


def test_raw_json_sheet_contains_json(sample_model_inputs):
    """Test that Raw JSON sheet contains JSON data."""
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Raw_JSON')
    
    # Check JSON column exists
    assert 'JSON' in df.columns
    
    # Check JSON content is present
    json_lines = df['JSON'].tolist()
    json_str = '\n'.join([str(line) for line in json_lines if pd.notna(line)])
    
    # Verify key fields are in JSON
    assert 'scenario_name' in json_str
    assert 'time_mode' in json_str
    assert 'revenue_streams' in json_str


def test_export_with_advanced_financing(sample_model_inputs):
    """Test export with advanced financing mode."""
    sample_model_inputs['mode'] = 'Advanced'
    sample_model_inputs['business_loan_amount'] = 75000.0
    sample_model_inputs['business_interest_rate'] = 0.055
    sample_model_inputs['business_amort_years'] = 7
    sample_model_inputs['real_estate_loan_amount'] = 200000.0
    sample_model_inputs['real_estate_interest_rate'] = 0.065
    sample_model_inputs['real_estate_amort_years'] = 20
    
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Scenario')
    
    # Check advanced financing parameters are present
    parameters = df['Parameter'].tolist()
    
    assert 'ADVANCED FINANCING' in parameters
    assert 'Business Loan Amount' in parameters
    assert 'Real Estate Loan Amount' in parameters


def test_export_with_multiple_revenue_streams(sample_model_inputs):
    """Test export with multiple revenue streams."""
    sample_model_inputs['revenue_streams'] = [
        {
            'name': 'Product Sales',
            'price': 100.0,
            'volume': 50.0,
            'growth_rate': 0.10,
            'cogs_override': None
        },
        {
            'name': 'Service Revenue',
            'price': 150.0,
            'volume': 30.0,
            'growth_rate': 0.15,
            'cogs_override': 0.20
        }
    ]
    
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Scenario')
    
    # Check revenue streams are present
    parameters = df['Parameter'].tolist()
    
    assert 'Stream 1: Product Sales' in parameters
    assert 'Stream 2: Service Revenue' in parameters


def test_export_with_multiple_payroll_roles(sample_model_inputs):
    """Test export with multiple payroll roles."""
    sample_model_inputs['payroll_roles'] = [
        {
            'name': 'Manager',
            'count': 1,
            'annual_salary': 50000.0,
            'payroll_tax_rate': 0.15,
            'category': 'indirect'
        },
        {
            'name': 'Sales Staff',
            'count': 3,
            'annual_salary': 35000.0,
            'payroll_tax_rate': 0.15,
            'category': 'direct'
        }
    ]
    
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Scenario')
    
    # Check payroll roles are present
    parameters = df['Parameter'].tolist()
    
    assert 'Role 1: Manager' in parameters
    assert 'Role 2: Sales Staff' in parameters


def test_export_without_dataframes_no_summary_sheet(sample_model_inputs):
    """Test that Summary sheet is not created when DataFrames are None."""
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Load workbook
    wb = load_workbook(BytesIO(excel_bytes))
    
    # Summary and statement sheets should not exist
    assert 'Summary' not in wb.sheetnames
    assert 'Income_Statement' not in wb.sheetnames
    assert 'Cash_Flow' not in wb.sheetnames


def test_export_file_is_valid_xlsx(sample_model_inputs):
    """Test that exported file is a valid Excel file."""
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Should be able to load as workbook without errors
    wb = load_workbook(BytesIO(excel_bytes))
    
    # Should have at least 2 sheets
    assert len(wb.sheetnames) >= 2


def test_export_with_owner_payroll_mode(sample_model_inputs):
    """Test export with owner compensation in payroll mode."""
    sample_model_inputs['owner_compensation'] = {
        'mode': 'payroll',
        'amount': 75000.0
    }
    
    excel_bytes = export_scenario_to_excel(sample_model_inputs)
    
    # Read Excel file
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name='Scenario')
    
    # Create lookup dictionary
    data_dict = dict(zip(df['Parameter'], df['Value']))
    
    # Verify owner compensation mode
    assert data_dict['Mode'] == 'Payroll'
    assert data_dict['Annual Amount'] == '$75,000.00'
